from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
import httpx
from datetime import datetime, timezone, timedelta
import io
import csv
import json
import math
import re

import bcrypt
import openpyxl
import xlsxwriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

def slugify_name(name):
    """Convert name to safe filename"""
    return re.sub(r'[^\w\-]', '_', name.strip())

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ MODELS ============

class MasjidCreate(BaseModel):
    name: str
    location_link: Optional[str] = ""
    waqth_chart_id: Optional[str] = ""
    adjustments: Optional[Dict[str, Any]] = {}
    district: Optional[str] = ""
    state: Optional[str] = ""
    masjid_address: Optional[str] = ""
    imam: Optional[str] = ""
    imam_number: Optional[str] = ""
    muaddin: Optional[str] = ""
    muaddin_number: Optional[str] = ""
    committee_members: Optional[List[Any]] = []
    remark: Optional[str] = ""
    mihrab_masjid_id: Optional[str] = ""

class WaqthChartCreate(BaseModel):
    name: str
    district: Optional[str] = ""
    state: Optional[str] = ""
    country: Optional[str] = ""
    prayer_times: Optional[List[Dict[str, Any]]] = []

class SalahConfigCreate(BaseModel):
    masjid_id: str
    chart_number: int = 1
    waqth_chart_id: str
    adjustments: Dict[str, Any] = {}

class GenerateSalahRequest(BaseModel):
    masjid_id: str
    chart_number: int = 1
    month: Optional[int] = None
    year: Optional[int] = None

# ============ AUTH HELPERS ============

async def get_current_user(request: Request):
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    session_doc = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return user_doc

# ============ AUTH ENDPOINTS ============

def _set_session_cookie(response: Response, token: str):
    response.set_cookie(
        key="session_token",
        value=token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )

async def _create_session(user_id: str, response: Response):
    session_token = uuid.uuid4().hex
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    _set_session_cookie(response, session_token)
    return session_token

@api_router.post("/auth/register")
async def register(request: Request, response: Response):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    password = body.get("password", "")
    name = body.get("name", "").strip()

    if not email or not password or not name:
        raise HTTPException(status_code=400, detail="Name, email and password are required")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    existing = await db.users.find_one({"email": email})
    if existing:
        if existing.get("password_hash"):
            raise HTTPException(status_code=400, detail="Email already registered")
        # Old Google OAuth account — add password and preserve all data
        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        await db.users.update_one({"email": email}, {"$set": {"password_hash": password_hash}})
        user_id = existing["user_id"]
    else:
        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": name,
            "password_hash": password_hash,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    token = await _create_session(user_id, response)
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return {**user_doc, "session_token": token}

@api_router.post("/auth/login")
async def login(request: Request, response: Response):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    password = body.get("password", "")

    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")

    user_doc = await db.users.find_one({"email": email})
    if not user_doc or not bcrypt.checkpw(password.encode("utf-8"), user_doc["password_hash"].encode("utf-8")):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    token = await _create_session(user_doc["user_id"], response)
    result = {k: v for k, v in user_doc.items() if k not in ("_id", "password_hash")}
    return {**result, "session_token": token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/", secure=True, samesite="none")
    return {"message": "Logged out"}

# ============ MASJID ENDPOINTS ============

@api_router.post("/masjids")
async def create_masjid(data: MasjidCreate, request: Request):
    user = await get_current_user(request)
    masjid_id = f"masjid_{uuid.uuid4().hex[:12]}"
    doc = {
        "masjid_id": masjid_id,
        "name": data.name,
        "location_link": data.location_link,
        "waqth_chart_id": data.waqth_chart_id,
        "adjustments": data.adjustments,
        "district": data.district,
        "state": data.state,
        "masjid_address": data.masjid_address,
        "imam": data.imam,
        "imam_number": data.imam_number,
        "muaddin": data.muaddin,
        "muaddin_number": data.muaddin_number,
        "committee_members": data.committee_members,
        "remark": data.remark,
        "mihrab_masjid_id": data.mihrab_masjid_id,
        "user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.masjids.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/masjids")
async def list_masjids(request: Request, search: Optional[str] = None):
    user = await get_current_user(request)
    query = {"user_id": user["user_id"]}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"mihrab_masjid_id": {"$regex": search, "$options": "i"}},
        ]
    masjids = await db.masjids.find(query, {"_id": 0}).to_list(1000)
    return masjids

@api_router.get("/masjids/{masjid_id}")
async def get_masjid(masjid_id: str, request: Request):
    user = await get_current_user(request)
    masjid = await db.masjids.find_one({"masjid_id": masjid_id, "user_id": user["user_id"]}, {"_id": 0})
    if not masjid:
        raise HTTPException(status_code=404, detail="Masjid not found")
    return masjid

@api_router.put("/masjids/{masjid_id}")
async def update_masjid(masjid_id: str, data: MasjidCreate, request: Request):
    user = await get_current_user(request)
    result = await db.masjids.update_one(
        {"masjid_id": masjid_id, "user_id": user["user_id"]},
        {"$set": {
            "name": data.name,
            "location_link": data.location_link,
            "waqth_chart_id": data.waqth_chart_id,
            "adjustments": data.adjustments,
            "district": data.district,
            "state": data.state,
            "masjid_address": data.masjid_address,
            "imam": data.imam,
            "imam_number": data.imam_number,
            "muaddin": data.muaddin,
            "muaddin_number": data.muaddin_number,
            "committee_members": data.committee_members,
            "remark": data.remark,
            "mihrab_masjid_id": data.mihrab_masjid_id,
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Masjid not found")
    masjid = await db.masjids.find_one({"masjid_id": masjid_id}, {"_id": 0})
    return masjid

@api_router.delete("/masjids/{masjid_id}")
async def delete_masjid(masjid_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.masjids.delete_one({"masjid_id": masjid_id, "user_id": user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Masjid not found")
    await db.salah_configs.delete_many({"masjid_id": masjid_id})
    return {"message": "Deleted"}

# ============ WAQTH CHART ENDPOINTS ============

@api_router.post("/waqth-charts")
async def create_waqth_chart(data: WaqthChartCreate, request: Request):
    user = await get_current_user(request)
    chart_id = f"chart_{uuid.uuid4().hex[:12]}"
    doc = {
        "chart_id": chart_id,
        "name": data.name,
        "district": data.district,
        "state": data.state,
        "country": data.country,
        "prayer_times": data.prayer_times,
        "user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.waqth_charts.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/waqth-charts")
async def list_waqth_charts(request: Request):
    user = await get_current_user(request)
    charts = await db.waqth_charts.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(1000)
    return charts

@api_router.get("/waqth-charts/{chart_id}")
async def get_waqth_chart(chart_id: str, request: Request):
    user = await get_current_user(request)
    chart = await db.waqth_charts.find_one({"chart_id": chart_id, "user_id": user["user_id"]}, {"_id": 0})
    if not chart:
        raise HTTPException(status_code=404, detail="Chart not found")
    return chart

@api_router.get("/waqth-charts/{chart_id}/columns")
async def get_waqth_chart_columns(chart_id: str, request: Request):
    """Get the column names of a waqth chart for mapping purposes"""
    user = await get_current_user(request)
    chart = await db.waqth_charts.find_one({"chart_id": chart_id, "user_id": user["user_id"]}, {"_id": 0})
    if not chart:
        raise HTTPException(status_code=404, detail="Chart not found")
    prayer_times = chart.get("prayer_times", [])
    if prayer_times:
        columns = list(prayer_times[0].keys())
    else:
        columns = []
    return {"columns": columns, "chart_name": chart.get("name", "")}

@api_router.delete("/waqth-charts/{chart_id}")
async def delete_waqth_chart(chart_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.waqth_charts.delete_one({"chart_id": chart_id, "user_id": user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Chart not found")
    return {"message": "Deleted"}

@api_router.post("/waqth-charts/upload")
async def upload_waqth_chart(
    request: Request,
    file: UploadFile = File(...),
    name: str = Form(...),
    district: str = Form(""),
    state: str = Form(""),
    country: str = Form("")
):
    user = await get_current_user(request)
    content = await file.read()
    prayer_times = []
    
    try:
        if file.filename.endswith('.csv'):
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                # Normalize keys to lowercase
                prayer_times.append({k.strip().lower(): v for k, v in row.items()})
        elif file.filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = "" if value is None else str(value)
                prayer_times.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV or Excel.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
    
    chart_id = f"chart_{uuid.uuid4().hex[:12]}"
    doc = {
        "chart_id": chart_id,
        "name": name,
        "district": district,
        "state": state,
        "country": country,
        "prayer_times": prayer_times,
        "user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.waqth_charts.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

# ============ SALAH CONFIG ENDPOINTS ============

@api_router.post("/salah-configs")
async def create_salah_config(data: SalahConfigCreate, request: Request):
    user = await get_current_user(request)
    config_id = f"config_{uuid.uuid4().hex[:12]}"
    doc = {
        "config_id": config_id,
        "masjid_id": data.masjid_id,
        "chart_number": data.chart_number,
        "waqth_chart_id": data.waqth_chart_id,
        "adjustments": data.adjustments,
        "user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    # Upsert - one config per masjid per chart_number
    await db.salah_configs.update_one(
        {"masjid_id": data.masjid_id, "chart_number": data.chart_number, "user_id": user["user_id"]},
        {"$set": doc},
        upsert=True
    )
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.get("/salah-configs/{masjid_id}")
async def get_salah_configs(masjid_id: str, request: Request):
    user = await get_current_user(request)
    configs = await db.salah_configs.find(
        {"masjid_id": masjid_id, "user_id": user["user_id"]}, {"_id": 0}
    ).to_list(10)
    return configs

# ============ SALAH GENERATION ============

def parse_time_str(time_str):
    """Parse various time formats into (hour, minute) tuple.
    Handles: '5:20', '05:20', '5.20', '5.2' (=5:20), '17.30', decimal day fraction (0.2166...)
    """
    if not time_str or str(time_str).strip() in ("", "-", "nan", "None"):
        return None
    s = str(time_str).strip()
    
    # Handle HH:MM format
    if ":" in s:
        parts = s.split(":")
        try:
            hour = int(parts[0])
            minute = int(parts[1].split()[0])  # handle "5:20 AM" edge case
            return (hour % 24, minute)
        except (ValueError, IndexError):
            return None
    
    # Handle period separator: "5.20" means 5:20, "17.05" means 17:05
    if "." in s:
        parts = s.split(".")
        try:
            hour_part = int(parts[0])
            min_part = parts[1]
            # If hour_part is 0 or 1 and looks like a day fraction (0.2166...)
            if hour_part == 0 and len(min_part) > 2:
                # Excel day fraction: multiply by 24 to get hours
                frac = float(s)
                total_minutes = int(round(frac * 24 * 60))
                return ((total_minutes // 60) % 24, total_minutes % 60)
            # Otherwise treat as H.MM format
            # "5.2" → ensure we interpret as 5:20 not 5:02
            # If minute part is 1 digit, multiply by 10 (5.2 = 5:20, 5.5 = 5:50)
            if len(min_part) == 1:
                minute = int(min_part) * 10
            else:
                minute = int(min_part[:2])
            return (hour_part % 24, min(minute, 59))
        except (ValueError, IndexError):
            return None
    
    # Try pure number (could be hour only like "5" meaning 5:00)
    try:
        val = float(s)
        if val < 1:
            # Day fraction
            total_minutes = int(round(val * 24 * 60))
            return ((total_minutes // 60) % 24, total_minutes % 60)
        elif val <= 24:
            return (int(val) % 24, 0)
        else:
            return None
    except ValueError:
        return None

def format_time(hour, minute):
    """Format hour and minute to HH:MM string"""
    return f"{hour:02d}:{minute:02d}"

def round_time(time_str, rule, custom_value=None):
    """Round a time string based on the rule.
    Rules: 'nearest_5', 'round_up_5', 'round_down_5', 'custom'
    custom_value: the exact minute value to use
    """
    parsed = parse_time_str(time_str)
    if parsed is None:
        return ""
    hour, minute = parsed
    
    if rule == "custom" and custom_value is not None:
        cv = int(custom_value)
        if cv == 0:
            return format_time(hour, minute)  # 0 = no change, preserve original
        return add_minutes_to_time(format_time(hour, minute), cv)
    elif rule == "nearest_5":
        rounded = round(minute / 5) * 5
        if rounded == 60:
            hour = (hour + 1) % 24
            rounded = 0
        return format_time(hour, rounded)
    elif rule == "round_up_5":
        rounded = math.ceil(minute / 5) * 5
        if rounded == 60:
            hour = (hour + 1) % 24
            rounded = 0
        return format_time(hour, rounded)
    elif rule == "round_down_5":
        rounded = math.floor(minute / 5) * 5
        return format_time(hour, rounded)
    else:
        # No rounding - just normalize format
        return format_time(hour, minute)

def add_minutes_to_time(time_str, minutes):
    """Add minutes to a time string HH:MM"""
    if not time_str or time_str == "":
        return ""
    parsed = parse_time_str(time_str)
    if parsed is None:
        return ""
    hour, minute = parsed
    total_minutes = hour * 60 + minute + int(minutes)
    new_hour = (total_minutes // 60) % 24
    new_minute = total_minutes % 60
    return format_time(new_hour, new_minute)

PRAYERS = ["fajr", "sunrise", "zuhr", "asr", "maghrib", "isha"]

# Common aliases for prayer column names in uploaded charts
PRAYER_ALIASES = {
    "fajr": ["fajr", "fajar", "fajir", "subh", "subuh", "fjar", "fazr"],
    "sunrise": ["sunrise", "shuruq", "shurooq", "ishraq", "tulu", "rise"],
    "zuhr": ["zuhr", "dhuhr", "zohr", "duhr", "luhr", "zuhar", "duhur", "zuhur", "zhur"],
    "asr": ["asr", "asar", "assr", "aasr", "aser", "assar", "acer"],
    "maghrib": ["maghrib", "magrib", "magreb", "maghreb", "magrib", "magrb", "mgrib"],
    "isha": ["isha", "ishaa", "esha", "ishai", "isha'", "ishak", "yatsi", "isya"],
}

# Prayers that are PM (if hour < 12, should add 12 for 24h conversion)
PM_PRAYERS = {"zuhr", "asr", "maghrib", "isha"}

def find_prayer_value(row_lower, prayer):
    """Find prayer time value using multiple aliases and partial matching"""
    aliases = PRAYER_ALIASES.get(prayer, [prayer])
    # Exact match first
    for alias in aliases:
        val = row_lower.get(alias, "")
        if val and str(val).strip() not in ("", "-", "nan", "None"):
            return str(val)
    # Partial match - check if any key contains the prayer name or alias
    for key in row_lower:
        for alias in aliases:
            if alias in key or key in alias:
                val = row_lower.get(key, "")
                if val and str(val).strip() not in ("", "-", "nan", "None"):
                    return str(val)
    return ""

def convert_12h_to_24h(hour, minute, prayer):
    """Convert 12-hour format to 24-hour for PM prayers.
    If Asr/Maghrib/Isha have hour < 12, they must be PM (add 12).
    Zuhr: if hour < 12 and hour > 0, add 12 (noon is 12, 1PM is 13).
    """
    if prayer in PM_PRAYERS:
        if prayer == "zuhr":
            # Zuhr is around noon: 11:30-14:00. If hour is 1-3, it's PM.
            if 0 < hour < 11:
                hour += 12
        else:
            # Asr (3-5pm), Maghrib (5-8pm), Isha (7-10pm)
            if hour < 12:
                hour += 12
    return hour, minute

@api_router.post("/generate-salah")
async def generate_salah(data: GenerateSalahRequest, request: Request):
    user = await get_current_user(request)
    
    config = await db.salah_configs.find_one(
        {"masjid_id": data.masjid_id, "chart_number": data.chart_number, "user_id": user["user_id"]},
        {"_id": 0}
    )
    if not config:
        raise HTTPException(status_code=404, detail="Salah config not found. Please configure adjustments first.")
    
    chart = await db.waqth_charts.find_one({"chart_id": config["waqth_chart_id"]}, {"_id": 0})
    if not chart:
        raise HTTPException(status_code=404, detail="Waqth chart not found")
    
    adjustments = config.get("adjustments", {})
    prayer_times = chart.get("prayer_times", [])
    
    generated = []
    for row in prayer_times:
        # Normalize all keys to lowercase for lookup
        row_lower = {k.lower().strip(): v for k, v in row.items()}
        gen_row = {"date": row_lower.get("date", "")}
        for prayer in PRAYERS:
            prayer_adj = adjustments.get(prayer, {})
            mode = prayer_adj.get("mode", "adjustment")  # "adjustment" or "fixed"
            
            if mode == "fixed":
                fixed_time = prayer_adj.get("fixed_time", "")
                gen_row[f"{prayer}_azan"] = fixed_time
            else:
                raw_time = find_prayer_value(row_lower, prayer)
                rounding_rule = prayer_adj.get("rounding", "nearest_5")
                custom_value = prayer_adj.get("custom_value", None)
                # Parse and convert 12h to 24h for PM prayers
                parsed = parse_time_str(raw_time)
                if parsed:
                    hour, minute = convert_12h_to_24h(parsed[0], parsed[1], prayer)
                    time_24h = format_time(hour, minute)
                    rounded = round_time(time_24h, rounding_rule, custom_value)
                    gen_row[f"{prayer}_azan"] = rounded if rounded else ""
                else:
                    gen_row[f"{prayer}_azan"] = ""
            
            # Iqamah calculation
            iqamah_offset = prayer_adj.get("iqamah_offset", 0)
            azan_val = gen_row[f"{prayer}_azan"]
            if iqamah_offset and azan_val:
                gen_row[f"{prayer}_iqamah"] = add_minutes_to_time(azan_val, iqamah_offset)
            else:
                gen_row[f"{prayer}_iqamah"] = ""
        
        # Jummah (Friday prayer) - always fixed time, manually entered
        jummah_adj = adjustments.get("jummah", {})
        jummah_time = jummah_adj.get("fixed_time", "")
        jummah_iqamah_offset = jummah_adj.get("iqamah_offset", 0)
        gen_row["jummah_azan"] = jummah_time
        if jummah_iqamah_offset and jummah_time:
            gen_row["jummah_iqamah"] = add_minutes_to_time(jummah_time, jummah_iqamah_offset)
        else:
            gen_row["jummah_iqamah"] = ""
        
        generated.append(gen_row)
    
    return {"generated": generated, "config": config}

# ============ EXPORT ENDPOINTS ============

@api_router.post("/export/csv")
async def export_csv(request: Request):
    await get_current_user(request)
    body = await request.json()
    data = body.get("data", [])
    masjid_name = slugify_name(body.get("masjid_name", "salah_times"))
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={masjid_name}_salah_times.csv"}
    )

@api_router.post("/export/excel")
async def export_excel(request: Request):
    await get_current_user(request)
    body = await request.json()
    data = body.get("data", [])
    masjid_name = slugify_name(body.get("masjid_name", "salah_times"))
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Salah Times')
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        for row_idx, row in enumerate(data, 1):
            for col, key in enumerate(headers):
                worksheet.write(row_idx, col, row.get(key, ""))
    workbook.close()
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={masjid_name}_salah_times.xlsx"}
    )

@api_router.post("/export/pdf")
async def export_pdf(request: Request):
    await get_current_user(request)
    body = await request.json()
    data = body.get("data", [])
    masjid_name_raw = body.get("masjid_name", "Salah Times")
    masjid_name = slugify_name(masjid_name_raw)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph(f"<b>{masjid_name_raw} - Salah Times</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    if data:
        headers = list(data[0].keys())
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])
        
        t = RLTable(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B5336')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F0')]),
        ]))
        elements.append(t)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={masjid_name}_salah_times.pdf"}
    )

# ============ DASHBOARD STATS ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    user = await get_current_user(request)
    user_id = user["user_id"]
    
    masjid_count = await db.masjids.count_documents({"user_id": user_id})
    chart_count = await db.waqth_charts.count_documents({"user_id": user_id})
    config_count = await db.salah_configs.count_documents({"user_id": user_id})
    
    recent_masjids = await db.masjids.find(
        {"user_id": user_id}, {"_id": 0}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "masjid_count": masjid_count,
        "chart_count": chart_count,
        "config_count": config_count,
        "recent_masjids": recent_masjids
    }

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
